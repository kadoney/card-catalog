#!/usr/bin/env python3
"""
Build the APF article-splitting plan: a workbook Keith corrects, then a splitter runs.

The Society's APF contents index already exists and is authoritative — Keith maintained
it by hand for years in the Joomla `sap_apf_master_index` table (recovered 2026-08-31 to
`card-catalog/data/apf-master-index.tsv`, 290 articles 2001-2024 with title, author and
PRINTED page). What is missing is the offset between a printed folio and the PDF's own
page index: one number per volume (2024 = +3).

So this does NOT re-read the contents. It joins the authoritative index to the page
ranges already computed by the 2026-09-04 text extraction (R2
`publications/pins-and-tales-text/apf-articles.jsonl`, 249 of 290 articles carrying
`offset`/`pdf_page_start`/`pdf_page_end`) and emits a workbook whose FIRST sheet is the
24 offsets — the only thing that actually needs a human.

⚠ Invent nothing: a volume whose offset could not be detected gets a BLANK offset and a
blank page range, never a guess. Four volumes (2001, 2003, 2008, 2010) refused detection
on 09-04 and are expected to come through empty.

Usage:
    python card-catalog/scripts/build_apf_split_plan.py \
        --articles-jsonl <apf-articles.jsonl> --pdf-dir <dir of apf-YYYY.pdf> \
        --out apf-split-plan.xlsx
"""
import argparse, collections, csv, json, os, re, sys

sys.stdout.reconfigure(encoding='utf-8')
HERE = os.path.dirname(os.path.abspath(__file__))
MASTER = os.path.join(HERE, '..', 'data', 'apf-master-index.tsv')


def norm(s):
    """Loose title key — OCR and TOC/printed-page disagreements are routine."""
    return re.sub(r'[^a-z0-9]+', ' ', (s or '').lower()).strip()


def load_master():
    with open(MASTER, encoding='utf-8') as fh:
        rows = list(csv.DictReader(fh, delimiter='\t'))
    out = []
    for r in rows:
        out.append({'year': int(r['year']), 'printed_page': int(r['page']),
                    'title': r['title'].strip(), 'author': r['author'].strip()})
    out.sort(key=lambda r: (r['year'], r['printed_page']))
    return out


def load_detected(path):
    det = {}
    with open(path, encoding='utf-8') as fh:
        for line in fh:
            if not line.strip():
                continue
            d = json.loads(line)
            det[(d['year'], d['page_start'])] = d
    return det


def pdf_page_counts(pdf_dir):
    from pypdf import PdfReader
    counts = {}
    if not pdf_dir or not os.path.isdir(pdf_dir):
        return counts
    for fn in os.listdir(pdf_dir):
        m = re.match(r'apf-(\d{4})\.pdf$', fn, re.I)
        if not m:
            continue
        try:
            counts[int(m.group(1))] = len(PdfReader(os.path.join(pdf_dir, fn)).pages)
        except Exception as e:
            print(f'  ! {fn}: {type(e).__name__} {e}')
    return counts


def detect_from_pdf(year, arts, pdf_dir):
    """Fallback for a volume the 09-04 extraction could not place.

    ⚠ A title's FIRST appearance is the table of contents, not the article — that is
    exactly what defeated the original run (it produced offsets like -67). Keith's call
    2026-09-05: take the SECOND occurrence. Returns (offset, votes, total, evidence) or
    (None, 0, n, '') — a plurality of one is still reported, with its count, so the
    workbook can say how thin the evidence is rather than presenting it as settled.
    """
    import collections
    path = os.path.join(pdf_dir or '', f'apf-{year}.pdf')
    if not os.path.isfile(path):
        return None, 0, len(arts), ''
    from pypdf import PdfReader
    try:
        txt = [norm(pg.extract_text() or '') for pg in PdfReader(path).pages]
    except Exception:
        return None, 0, len(arts), ''
    votes, ev = collections.Counter(), collections.defaultdict(list)
    for a in arts:
        hits = []
        for L in (34, 26, 20):
            key = norm(a['title'])[:L]
            if len(key) < 12:
                continue
            hits = [i + 1 for i, t in enumerate(txt) if key in t]
            if len(hits) >= 2:
                break
        if len(hits) >= 2:
            off = hits[1] - a['printed_page']
            votes[off] += 1
            ev[off].append(f"printed p{a['printed_page']} -> PDF p{hits[1]} ({a['title'][:34]})")
    if not votes:
        return None, 0, len(arts), ''
    off, n = votes.most_common(1)[0]
    return off, n, len(arts), '; '.join(ev[off][:2])


def build(master, det, counts, pdf_dir=None):
    """Per-volume offset by majority vote over the articles that were located."""
    votes = collections.defaultdict(list)
    for (year, page), d in det.items():
        if d.get('offset') is not None:
            votes[year].append(d['offset'])

    volumes = {}
    for year in sorted({r['year'] for r in master}):
        v = votes.get(year, [])
        n_articles = sum(1 for r in master if r['year'] == year)
        if v:
            tally = collections.Counter(v)
            off, agree = tally.most_common(1)[0]
        else:
            off, agree, tally = None, 0, collections.Counter()
        candidate, cand_n, cand_ev = None, 0, ''
        if off is None:
            rows_y = [r for r in master if r['year'] == year]
            candidate, cand_n, _, cand_ev = detect_from_pdf(year, rows_y, pdf_dir)
            # ⚠ A candidate that would push the last article past the end of the PDF is
            # impossible, not merely weak. Reject rather than hand it over to be checked.
            pp = counts.get(year)
            if candidate is not None and pp:
                last = max(r['printed_page'] for r in rows_y)
                if last + candidate > pp or min(r['printed_page'] for r in rows_y) + candidate < 1:
                    cand_ev = (f'rejected {candidate:+d}: it would place printed p{last} at '
                               f'PDF p{last + candidate} in a {pp}-page file')
                    candidate, cand_n = None, 0
        volumes[year] = {
            'year': year, 'pdf_pages': counts.get(year), 'articles': n_articles,
            'located': len(v), 'offset': off, 'agree': agree,
            'unanimous': bool(v) and len(tally) == 1,
            'spread': ', '.join(f'{o}x{c}' for o, c in sorted(tally.items())) if len(tally) > 1 else '',
            'candidate': candidate, 'cand_n': cand_n, 'cand_ev': cand_ev,
        }

    # articles, in printed order, with end page = next article's start - 1
    arts = []
    by_year = collections.defaultdict(list)
    for r in master:
        by_year[r['year']].append(r)
    for year, rows in sorted(by_year.items()):
        vol = volumes[year]
        off = vol['offset'] if vol['offset'] is not None else vol.get('candidate')
        off_is_candidate = vol['offset'] is None and vol.get('candidate') is not None
        for i, r in enumerate(rows):
            d = det.get((year, r['printed_page']))
            if d:
                start, end, src = d['pdf_page_start'], d['pdf_page_end'], 'located'
                verified = norm(d['title'])[:24] == norm(r['title'])[:24]
            elif off is not None:
                start = r['printed_page'] + off
                nxt = rows[i + 1]['printed_page'] + off if i + 1 < len(rows) else None
                end = (nxt - 1) if nxt else vol['pdf_pages']
                src = ('CANDIDATE offset — please verify' if off_is_candidate
                       else 'derived from volume offset')
                verified = False
            else:
                start = end = None
                src, verified = 'NEEDS OFFSET', False
            arts.append({'year': year, 'seq': i + 1, 'printed_page': r['printed_page'],
                         'title': r['title'], 'author': r['author'], 'offset': off,
                         'pdf_start': start, 'pdf_end': end, 'source': src,
                         'title_verified': verified})
    return volumes, arts


HDR = {'fill': 'FFE8E0D4', 'font': 'FF3B2F2A'}
ASK = 'FFFFF2CC'   # columns Keith fills
WARN = 'FFF4CCCC'  # rows that cannot proceed


def emit(volumes, arts, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = PatternFill('solid', fgColor=HDR['fill'])
    hfont = Font(bold=True, color=HDR['font'])
    ask = PatternFill('solid', fgColor=ASK)
    warn = PatternFill('solid', fgColor=WARN)

    def sheet(ws, cols, widths):
        ws.append(cols)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in ws[1]:
            c.fill, c.font = hf, hfont
            c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.freeze_panes = 'A2'

    # --- Start here ---
    ws = wb.create_sheet('Notes')
    ws.column_dimensions['A'].width = 110
    lines = [
        ('APF article splitting — what I need from you', True),
        ('', False),
        ('The contents are already right. Title, author and printed page for all 290 articles come', False),
        ('from the APF master index you maintained by hand for years — recovered from the retired', False),
        ('Joomla site. You do not need to re-read a single table of contents.', False),
        ('', False),
        ('What is missing is one number per volume: the OFFSET between the printed page number and', False),
        ('the PDF\'s own page index. In the 2024 volume, printed page 6 is PDF page 9, so the offset', False),
        ('is +3. Get that number right and all of that volume\'s articles fall out correctly.', False),
        ('', False),
        ('THE JOB — the "Volumes" sheet, column "YOUR OFFSET".', True),
        ('  • Where I detected an offset and every article agreed, leave it alone.', False),
        ('  • Where the articles DISAGREED, the spread column shows what I saw. Check one article.', False),
        ('  • Where the offset is blank, I could not detect it at all and have guessed nothing.', False),
        ('    Open that volume, find the first article, and note which PDF page it starts on.', False),
        ('    Offset = PDF page − printed page. Four volumes are expected to be blank.', False),
        ('', False),
        ('The "Articles" sheet is there so you can see the consequence and override any single row.', False),
        ('You should not need to touch it much. Rows I could not place are shaded.', False),
        ('', False),
        ('Shaded columns are yours to fill. Everything else is mine — overwrite it if it is wrong.', False),
        ('Hand the file back and I will split and name the PDFs.', False),
    ]
    for text, bold in lines:
        ws.append([text])
        if bold:
            ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)

    # --- Volumes ---
    ws = wb.create_sheet('Volumes')
    sheet(ws, ['Year', 'PDF pages', 'Articles', 'Articles I located', 'Offset I detected',
               'Articles agreeing', 'Disagreement', 'YOUR OFFSET', 'Note'],
          [8, 11, 10, 11, 12, 11, 18, 12, 52])
    for y in sorted(volumes):
        v = volumes[y]
        if v['offset'] is None and v.get('candidate') is not None:
            note = (f"NOT DETECTED. Candidate {v['candidate']:+d} from {v['cand_n']} of "
                    f"{v['articles']} articles (2nd occurrence of the title, since the 1st "
                    f"is the contents page). PLEASE VERIFY — evidence: {v['cand_ev']}")
        elif v['offset'] is None:
            note = 'NOT DETECTED and no candidate. Open the volume: offset = PDF page of an article − its printed page.'
        elif not v['unanimous']:
            note = 'Articles disagreed; most common value shown. Worth one spot-check.'
        elif v['located'] < v['articles']:
            note = f"{v['articles'] - v['located']} article(s) not individually located; offset applied to them."
        else:
            note = 'All articles located and agreed.'
        shown = v['offset'] if v['offset'] is not None else v.get('candidate')
        ws.append([y, v['pdf_pages'], v['articles'], v['located'], v['offset'],
                   v['agree'], v['spread'], shown, note])
        r = ws.max_row
        ws.cell(r, 8).fill = ask
        if v['offset'] is None:
            for c in range(1, 10):
                ws.cell(r, c).fill = warn
            ws.cell(r, 8).fill = ask

    # --- Articles ---
    ws = wb.active
    ws.title = 'Articles'
    sheet(ws, ['Year', '#', 'Title', 'Author', 'Printed page', 'PDF start', 'PDF stop',
               'Confidence', 'YOUR start', 'YOUR stop', 'Note'],
          [7, 4, 50, 24, 12, 10, 10, 26, 11, 11, 26])
    for a in arts:
        conf = {'located': 'found on the page',
                'derived from volume offset': 'from volume offset',
                'CANDIDATE offset — please verify': 'CANDIDATE — verify',
                'NEEDS OFFSET': 'NO OFFSET — blank'}.get(a['source'], a['source'])
        ws.append([a['year'], a['seq'], a['title'], a['author'], a['printed_page'],
                   a['pdf_start'], a['pdf_end'], conf, a['pdf_start'], a['pdf_end'], ''])
        r = ws.max_row
        for c in (9, 10):
            ws.cell(r, c).fill = ask
        if a['pdf_start'] is None or 'CANDIDATE' in conf:
            for c in range(1, 12):
                ws.cell(r, c).fill = warn
            for c in (9, 10):
                ws.cell(r, c).fill = ask
    ws.auto_filter.ref = ws.dimensions

    wb._sheets = [wb['Articles'], wb['Volumes'], wb['Notes']]
    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--articles-jsonl', required=True)
    ap.add_argument('--pdf-dir')
    ap.add_argument('--out', default='apf-split-plan.xlsx')
    a = ap.parse_args()

    master = load_master()
    det = load_detected(a.articles_jsonl)
    counts = pdf_page_counts(a.pdf_dir)
    volumes, arts = build(master, det, counts, a.pdf_dir)

    placed = sum(1 for x in arts if x['pdf_start'] is not None)
    print(f'master index : {len(master)} articles, {len(volumes)} volumes')
    print(f'located      : {len(det)} articles carried a detected page range')
    print(f'placed       : {placed}/{len(arts)}  ({len(arts) - placed} need an offset)')
    print(f'pdf pages    : {len(counts)}/{len(volumes)} volumes measured')
    print()
    print(f"{'year':>5} {'off':>4} {'loc':>4} {'art':>4} {'pp':>4}  note")
    for y in sorted(volumes):
        v = volumes[y]
        flag = (f"CANDIDATE {v['candidate']:+d} ({v['cand_n']}/{v['articles']}) — verify"
                if v['offset'] is None and v.get('candidate') is not None
                else 'BLANK — no candidate') if v['offset'] is None else (
            f"disagreed: {v['spread']}" if not v['unanimous'] else '')
        print(f"{y:>5} {str(v['offset']):>4} {v['located']:>4} {v['articles']:>4} "
              f"{str(v['pdf_pages']):>4}  {flag}")
    emit(volumes, arts, a.out)
    print(f'\nwrote {a.out}')


if __name__ == '__main__':
    main()
