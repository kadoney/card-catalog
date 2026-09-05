#!/usr/bin/env python3
"""
Cut each APF volume into per-article PDFs, from the page ranges Keith corrected.

Reads the workbook produced by build_apf_split_plan.py AFTER Keith has been through it,
and uses the 'YOUR start'/'YOUR stop' columns — never my own 'PDF start'/'PDF stop'.
Those two columns are the whole point: 2008's offset genuinely DRIFTS within the volume
(+2, then +4, then +6), so no single per-volume offset can be right, and a row-level
range is the only representation that survives contact with the real files.

Naming follows the convention the lost 2026-era set used, recorded in
memory/project-apf-ingestion.md:

    APF_{year}_{NN} - {title} - {author}.pdf

⚠ Output is PAID CONTENT (APF is a separate product from membership — the volume is held
out of the search index by sapfm-embedder's RESTRICTED_CARD_SQL). Write it outside every
git repo and do not upload it anywhere without that being a decision.

Usage:
    python card-catalog/scripts/split_apf_articles.py \
        --plan apf-split-plan-V2.xlsx --pdf-dir ~/Downloads/apf-quality \
        --out ~/Downloads/apf-articles [--dry-run]
"""
import argparse, os, re, sys

sys.stdout.reconfigure(encoding='utf-8')

BAD = re.compile(r'[\/:*?"<>|]')          # illegal in Windows filenames
WS = re.compile(r'\s+')


def safe(s, limit=90):
    """Filename-safe, but keep it readable — these get browsed by people."""
    s = BAD.sub('', (s or '').replace('&', 'and'))
    s = WS.sub(' ', s).strip(' .')
    return s[:limit].strip(' .')


def load_plan(path):
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True)['Articles']
    hdr = [c.value for c in ws[1]]
    need = {'Year', '#', 'Title', 'Author', 'YOUR start', 'YOUR stop'}
    missing = need - set(hdr)
    if missing:
        sys.exit(f'plan is missing column(s): {sorted(missing)}')
    out = []
    for r in ws.iter_rows(min_row=2):
        d = dict(zip(hdr, [c.value for c in r]))
        out.append(d)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--plan', required=True)
    ap.add_argument('--pdf-dir', required=True)
    ap.add_argument('--out', required=True)
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()

    from pypdf import PdfReader, PdfWriter
    rows = load_plan(a.plan)
    pdf_dir = os.path.expanduser(a.pdf_dir)
    outdir = os.path.expanduser(a.out)

    by_year = {}
    for r in rows:
        by_year.setdefault(r['Year'], []).append(r)

    written = skipped = 0
    problems = []
    for year in sorted(by_year):
        src = os.path.join(pdf_dir, f'apf-{year}.pdf')
        if not os.path.isfile(src):
            problems.append(f'{year}: no source PDF at {src}')
            continue
        reader = PdfReader(src)
        n = len(reader.pages)
        ydir = os.path.join(outdir, str(year))
        if not a.dry_run:
            os.makedirs(ydir, exist_ok=True)

        for r in sorted(by_year[year], key=lambda x: x['#']):
            start, stop = r['YOUR start'], r['YOUR stop']
            if not isinstance(start, int) or not isinstance(stop, int):
                problems.append(f"{year} #{r['#']}: no page range — skipped")
                skipped += 1
                continue
            if not (1 <= start <= stop <= n):
                problems.append(f"{year} #{r['#']}: range {start}-{stop} outside 1-{n} — skipped")
                skipped += 1
                continue
            name = f"APF_{year}_{int(r['#']):02d} - {safe(r['Title'])} - {safe(r['Author'], 50)}.pdf"
            dest = os.path.join(ydir, name)
            if a.dry_run:
                print(f'  would write {year}/{name}  (pp {start}-{stop})')
            else:
                w = PdfWriter()
                for p in range(start - 1, stop):
                    w.add_page(reader.pages[p])
                with open(dest, 'wb') as fh:
                    w.write(fh)
            written += 1

    print(f'\n{"would write" if a.dry_run else "wrote"}: {written} article PDFs'
          f'{"" if not skipped else f", skipped {skipped}"}')
    if problems:
        print('\nproblems:')
        for p in problems:
            print('  ', p)
    if not a.dry_run:
        print(f'\noutput: {outdir}')


if __name__ == '__main__':
    main()
