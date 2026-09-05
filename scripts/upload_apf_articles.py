#!/usr/bin/env python3
"""
Upload the per-article APF PDFs to R2 and write a manifest tying each to its card.

⚠ PAID CONTENT. APF is a separate product from membership: `sapfm-embedder`'s
RESTRICTED_CARD_SQL holds card_type='apf-issue' out of the search index, and the
volume PDFs are reachable only through sapfm-catalog-api's JWT-gated /publications/*
route. These articles go to the SAME bucket and prefix family, so they inherit exactly
that protection — public r2.dev access on `publications` is disabled and must stay so.

Uploading does NOT grant anyone access to anything. It stores the objects and records
the mapping. Wiring library_cards.view_url/download_url to them is a SEPARATE decision
(a deliberate one — see project-apf-digital-access-ownership-mirror) and this script
does not touch D1.

Keys:  apf/{year}/articles/apf-{year}-{NN}-{slug}.pdf

Usage:
    python card-catalog/scripts/upload_apf_articles.py \
        --dir ~/Downloads/apf-articles --manifest apf-article-manifest.json [--execute]
"""
import argparse, hashlib, json, os, re, sys, urllib.request

sys.stdout.reconfigure(encoding='utf-8')
BUCKET = 'publications'


def slug(s, limit=60):
    s = re.sub(r'[^a-z0-9]+', '-', (s or '').lower()).strip('-')
    return s[:limit].strip('-')


def client():
    import boto3
    tok = open(os.path.expanduser('~/.sapfm/claude-cf-token')).read().strip()
    acct = os.environ['CLOUDFLARE_ACCOUNT_ID']
    tid = json.load(urllib.request.urlopen(urllib.request.Request(
        f'https://api.cloudflare.com/client/v4/accounts/{acct}/tokens/verify',
        headers={'Authorization': f'Bearer {tok}'})))['result']['id']
    return boto3.client('s3', endpoint_url=f'https://{acct}.r2.cloudflarestorage.com',
                        aws_access_key_id=tid,
                        aws_secret_access_key=hashlib.sha256(tok.encode()).hexdigest(),
                        region_name='auto')


def d1_rows(path):
    d = json.load(open(path, encoding='utf-8'))
    rs = d[0]['results'] if isinstance(d, list) else d['result'][0]['results']
    return {(r['year'], r['page_start']): r for r in rs}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dir', required=True)
    ap.add_argument('--plan', required=True)
    ap.add_argument('--d1-rows', required=True, help='JSON from wrangler d1 execute')
    ap.add_argument('--manifest', default='apf-article-manifest.json')
    ap.add_argument('--execute', action='store_true')
    a = ap.parse_args()

    import openpyxl
    ws = openpyxl.load_workbook(os.path.expanduser(a.plan), data_only=True)['Articles']
    hdr = [c.value for c in ws[1]]
    plan = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
    cards = d1_rows(os.path.expanduser(a.d1_rows))
    root = os.path.expanduser(a.dir)

    entries, unmatched, missing = [], [], []
    for p in plan:
        card = cards.get((p['Year'], p['Printed page']))
        if not card:
            unmatched.append(f"{p['Year']} p{p['Printed page']} {p['Title'][:40]}")
            continue
        # the local filename is whatever split_apf_articles.py wrote
        ydir = os.path.join(root, str(p['Year']))
        pref = f"APF_{p['Year']}_{int(p['#']):02d} - "
        hit = [f for f in os.listdir(ydir)] if os.path.isdir(ydir) else []
        hit = [f for f in hit if f.startswith(pref)]
        if not hit:
            missing.append(pref)
            continue
        local = os.path.join(ydir, hit[0])
        key = (f"apf/{p['Year']}/articles/"
               f"apf-{p['Year']}-{int(p['#']):02d}-{slug(p['Title'])}.pdf")
        entries.append({'card_id': card['id'], 'parent_id': card['parent_id'],
                        'year': p['Year'], 'seq': int(p['#']),
                        'title': p['Title'], 'author': p['Author'],
                        'printed_page': p['Printed page'],
                        'pdf_start': p['YOUR start'], 'pdf_stop': p['YOUR stop'],
                        'key': key, 'bytes': os.path.getsize(local),
                        'local': local})

    print(f'matched {len(entries)}/{len(plan)} articles to D1 cards')
    if unmatched:
        print(f'  ⚠ no D1 card for {len(unmatched)}:'); [print('   ', u) for u in unmatched[:8]]
    if missing:
        print(f'  ⚠ no local PDF for {len(missing)}:'); [print('   ', m) for m in missing[:8]]
    if unmatched or missing:
        sys.exit('refusing to proceed with an incomplete mapping')

    total = sum(e['bytes'] for e in entries)
    print(f'total {total/1e9:.2f} GB across {len(entries)} objects')
    if not a.execute:
        for e in entries[:3]:
            print('  would put', e['key'])
        print('\ndry run — pass --execute to upload')
    else:
        s3 = client()
        for i, e in enumerate(entries, 1):
            s3.upload_file(e['local'], BUCKET, e['key'],
                           ExtraArgs={'ContentType': 'application/pdf'})
            if i % 25 == 0 or i == len(entries):
                print(f'  {i}/{len(entries)}', flush=True)
    for e in entries:
        e.pop('local', None)
    with open(os.path.expanduser(a.manifest), 'w', encoding='utf-8') as fh:
        json.dump(entries, fh, indent=1, ensure_ascii=False)
    print(f'manifest -> {a.manifest}')


if __name__ == '__main__':
    main()
